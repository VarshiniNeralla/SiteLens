from __future__ import annotations

from io import BytesIO
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels
from PIL import Image as PILImage

from app.domain import ObservationRecord
from app.logging_config import get_logger
from app.services import cloudinary_service
from app.services.observation_text import (
    effective_observation_narrative,
    effective_recommendation_narrative,
)

logger = get_logger(__name__)

_THUMB_MAX_W = 260
_THUMB_MAX_H = 195
_DATA_ROW_HEIGHT_PT = 132
_PREVIEW_COL_WIDTH_CHARS = 36
_PREVIEW_CELL_PADDING_PX = 8


def _d(v: object) -> str:
    if hasattr(v, "isoformat") and callable(getattr(v, "isoformat")):
        iso = getattr(v, "isoformat")()
        return iso if iso else "—"
    s = str(v).strip() if v is not None else ""
    return s if s else "—"


def _thumbnail_url(obs: ObservationRecord) -> str | None:
    pid = obs.cloudinary_public_id
    if pid and cloudinary_service.cloudinary_enabled():
        try:
            return cloudinary_service.build_excel_thumb_url(pid)
        except Exception as e:  # noqa: BLE001
            logger.debug("excel thumb transformation failed: %s", e)
    ref = ((obs.cloudinary_secure_url or obs.image_path or "").strip())
    return ref if ref.startswith(("http://", "https://")) else None


def _hyperlink_target(obs: ObservationRecord) -> str | None:
    u = ((obs.cloudinary_secure_url or obs.image_path or "").strip())
    return u if u.startswith(("http://", "https://")) else None


def _download_image_bytes(url: str) -> bytes | None:
    try:
        with httpx.Client(timeout=55.0, follow_redirects=True) as client:
            r = client.get(url)
            r.raise_for_status()
            return r.content
    except httpx.HTTPError as e:
        logger.warning("Excel image download failed (%s…): %s", url[:64], e)
        return None


def _bytes_to_xl_image(blob: bytes) -> XLImage | None:
    try:
        im = PILImage.open(BytesIO(blob))
        try:
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA")
            im.thumbnail((_THUMB_MAX_W, _THUMB_MAX_H), PILImage.Resampling.LANCZOS)
            out = BytesIO()
            im.save(out, format="PNG")
            out.seek(0)
            return XLImage(out)
        finally:
            im.close()
    except Exception as e:  # noqa: BLE001
        logger.warning("Excel thumbnail decode failed: %s", e)
        return None


def _column_width_to_pixels(width_chars: float | None) -> int:
    # Excel width units approximation used by openpyxl ecosystem.
    w = 8.43 if width_chars is None else float(width_chars)
    if w <= 0:
        return 64
    return int(w * 7 + 5)


def _row_height_to_pixels(height_points: float | None) -> int:
    pt = 15 if height_points is None else float(height_points)
    return int(points_to_pixels(pt))


def _anchor_image_to_cell(
    *,
    ws: object,
    img: XLImage,
    row_idx: int,
    col_idx: int,
    img_w: int,
    img_h: int,
) -> None:
    cell_w = _column_width_to_pixels(ws.column_dimensions[get_column_letter(col_idx)].width)
    cell_h = _row_height_to_pixels(ws.row_dimensions[row_idx].height)
    avail_w = max(24, cell_w - (_PREVIEW_CELL_PADDING_PX * 2))
    avail_h = max(24, cell_h - (_PREVIEW_CELL_PADDING_PX * 2))
    scale = min(avail_w / max(1, img_w), avail_h / max(1, img_h), 1.0)
    draw_w = max(20, int(round(img_w * scale)))
    draw_h = max(20, int(round(img_h * scale)))
    off_x = max(_PREVIEW_CELL_PADDING_PX, int((cell_w - draw_w) / 2))
    off_y = max(_PREVIEW_CELL_PADDING_PX, int((cell_h - draw_h) / 2))

    img.width = draw_w
    img.height = draw_h
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=col_idx - 1,
            row=row_idx - 1,
            colOff=pixels_to_EMU(off_x),
            rowOff=pixels_to_EMU(off_y),
        ),
        ext=XDRPositiveSize2D(pixels_to_EMU(draw_w), pixels_to_EMU(draw_h)),
    )


def build_quality_observation_xlsx(
    *,
    observations: list[ObservationRecord],
    output_path: Path,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "SiteLens Observations"

    headers = [
        "Observation ID",
        "Site name",
        "Tower",
        "Floor",
        "Flat",
        "Room",
        "Observation type",
        "Severity",
        "Site visit date",
        "Slab casting date",
        "Inspection status",
        "3rd-party status",
        "Recorded (UTC)",
        "Report observation text",
        "AI-generated observation",
        "Manual observation",
        "AI recommendation",
        "AI drafting status",
        "Image preview",
        "Image URL",
    ]

    ws.append(headers)

    preview_col_idx = headers.index("Image preview") + 1
    link_col_idx = headers.index("Image URL") + 1

    header_fill = PatternFill(fill_type="solid", fgColor="EAEAEE")
    alt_fill = PatternFill(fill_type="solid", fgColor="FAFBFC")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    thin_side = Side(border_style="thin", color="D7D9DE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    hdr_font = Font(name="Calibri", size=11, bold=True, color="1D1D1F")
    body_font = Font(name="Calibri", size=10, color="1D1D1F")
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    hdr_row = ws[1]
    for cell in hdr_row:
        cell.font = hdr_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    widths = [
        12,
        22,
        10,
        10,
        12,
        18,
        22,
        12,
        14,
        14,
        20,
        20,
        18,
        48,
        48,
        36,
        36,
        16,
        _PREVIEW_COL_WIDTH_CHARS,
        28,
    ]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for idx, obs in enumerate(observations, start=2):
        row_alt = idx % 2 == 0
        row_bg = alt_fill if row_alt else white_fill
        ws.row_dimensions[idx].height = _DATA_ROW_HEIGHT_PT

        ws.cell(row=idx, column=1).value = obs.id
        ws.cell(row=idx, column=2).value = obs.project_name
        ws.cell(row=idx, column=3).value = obs.tower
        ws.cell(row=idx, column=4).value = obs.floor
        ws.cell(row=idx, column=5).value = obs.flat
        ws.cell(row=idx, column=6).value = obs.room
        ws.cell(row=idx, column=7).value = obs.observation_type
        ws.cell(row=idx, column=8).value = obs.severity
        ws.cell(row=idx, column=9).value = _d(obs.site_visit_date)
        ws.cell(row=idx, column=10).value = _d(obs.slab_casting_date)
        ws.cell(row=idx, column=11).value = obs.inspection_status or "—"
        ws.cell(row=idx, column=12).value = obs.third_party_status or "—"
        ws.cell(row=idx, column=13).value = _d(obs.created_at)
        ws.cell(row=idx, column=14).value = effective_observation_narrative(obs) or "—"
        ai_status = (obs.ai_status or "").strip().lower() or "unavailable"
        ws.cell(row=idx, column=15).value = (obs.generated_observation or "").strip() or (
            "AI currently unavailable" if ai_status in {"unavailable", "failed"} else "—"
        )
        ws.cell(row=idx, column=16).value = (obs.manually_written_observation or "").strip() or "—"
        ws.cell(row=idx, column=17).value = effective_recommendation_narrative(obs) or (
            "Recommendation unavailable" if ai_status in {"unavailable", "failed"} else "—"
        )
        ws.cell(row=idx, column=18).value = ai_status

        tgt = _hyperlink_target(obs)
        link_cell = ws.cell(row=idx, column=link_col_idx)
        if tgt:
            link_cell.hyperlink = tgt
            link_cell.value = "Open full image"
            link_cell.font = link_font
        else:
            link_cell.value = "(local file — paste path from record)"
            link_cell.font = body_font

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=idx, column=c)
            if c != preview_col_idx:
                cell.fill = row_bg
            cell.border = border
            if c == preview_col_idx:
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=False)
            else:
                cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            if c not in (link_col_idx, preview_col_idx):
                cell.font = body_font

        ws.cell(row=idx, column=preview_col_idx).value = ""

        turl = _thumbnail_url(obs)
        if turl:
            blob = _download_image_bytes(turl)
            if blob:
                xli = _bytes_to_xl_image(blob)
                if xli is not None:
                    img_w = int(getattr(xli, "width", _THUMB_MAX_W) or _THUMB_MAX_W)
                    img_h = int(getattr(xli, "height", _THUMB_MAX_H) or _THUMB_MAX_H)
                    _anchor_image_to_cell(
                        ws=ws,
                        img=xli,
                        row_idx=idx,
                        col_idx=preview_col_idx,
                        img_w=img_w,
                        img_h=img_h,
                    )
                    ws.add_image(xli)
                else:
                    ws.cell(row=idx, column=preview_col_idx).value = "Preview unavailable"
            else:
                ws.cell(row=idx, column=preview_col_idx).value = "Preview unavailable"
        else:
            ws.cell(row=idx, column=preview_col_idx).value = "Preview unavailable"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
