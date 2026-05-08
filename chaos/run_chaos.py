from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import subprocess
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from PIL import Image
from pptx import Presentation


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


@dataclass
class ScenarioResult:
    name: str
    passed: bool
    duration_ms: int
    details: dict[str, Any]
    error: str | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "passed": self.passed,
            "duration_ms": self.duration_ms,
            "details": self.details,
            "error": self.error,
        }


class ChaosRunner:
    def __init__(self, base_url: str, out_dir: Path, backend_cmd: str | None = None) -> None:
        self.base_url = base_url.rstrip("/")
        self.out_dir = out_dir
        self.out_dir.mkdir(parents=True, exist_ok=True)
        self.client = httpx.Client(base_url=self.base_url, timeout=90.0)
        self.backend_cmd = backend_cmd
        self.backend_proc: subprocess.Popen[str] | None = None

    def close(self) -> None:
        self.client.close()
        if self.backend_proc and self.backend_proc.poll() is None:
            self.backend_proc.kill()

    def _api(self, method: str, path: str, **kwargs: Any) -> httpx.Response:
        r = self.client.request(method, path, **kwargs)
        return r

    def _json(self, method: str, path: str, **kwargs: Any) -> dict[str, Any]:
        r = self._api(method, path, **kwargs)
        r.raise_for_status()
        return r.json()

    def _wait_backend(self, timeout_s: float = 30.0) -> bool:
        end = time.time() + timeout_s
        while time.time() < end:
            try:
                r = self._api("GET", "/healthz")
                if r.status_code == 200:
                    return True
            except Exception:
                pass
            time.sleep(0.5)
        return False

    def _start_backend(self) -> None:
        if not self.backend_cmd:
            return
        if self.backend_proc and self.backend_proc.poll() is None:
            return
        self.backend_proc = subprocess.Popen(  # noqa: S603
            self.backend_cmd,
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        if not self._wait_backend():
            raise RuntimeError("Backend did not become healthy after start")

    def _restart_backend(self) -> None:
        if not self.backend_cmd:
            raise RuntimeError("backend_cmd not provided for restart scenario")
        if self.backend_proc and self.backend_proc.poll() is None:
            self.backend_proc.kill()
            self.backend_proc.wait(timeout=10)
        self.backend_proc = None
        self._start_backend()

    def _fault_reset(self) -> None:
        self._json("POST", "/api/ops/faults/reset")

    def _set_fault(self, target: str, enabled: bool, mode: str = "none", latency_ms: int = 0) -> None:
        self._json(
            "POST",
            "/api/ops/faults",
            json={
                "target": target,
                "enabled": enabled,
                "mode": mode,
                "latency_ms": latency_ms,
            },
        )

    def _make_png_bytes(self, w: int = 32, h: int = 32, color: tuple[int, int, int] = (220, 12, 12)) -> bytes:
        buf = io.BytesIO()
        Image.new("RGB", (w, h), color).save(buf, format="PNG")
        return buf.getvalue()

    def _resumable_upload(self, payload: bytes, filename: str = "chaos.png") -> dict[str, Any]:
        checksum = hashlib.sha256(payload).hexdigest()
        sess = self._json(
            "POST",
            "/api/upload/sessions",
            json={
                "filename": filename,
                "content_type": "image/png",
                "total_size": len(payload),
                "checksum_sha256": checksum,
            },
        )
        sid = sess["session_id"]
        chunk = max(64, int(sess.get("chunk_size", 512 * 1024)))
        # Simulate interruption: upload first half, query status, then resume.
        half = max(1, len(payload) // 2)
        off = 0
        while off < half:
            nxt = min(half, off + chunk)
            r = self._api(
                "PUT",
                f"/api/upload/sessions/{sid}/chunk",
                headers={"X-Chunk-Offset": str(off), "Content-Type": "application/octet-stream"},
                content=payload[off:nxt],
            )
            r.raise_for_status()
            off = r.json()["uploaded_bytes"]
        # "disconnect" period
        time.sleep(0.3)
        status = self._json("GET", f"/api/upload/sessions/{sid}")
        off = int(status["uploaded_bytes"])
        while off < len(payload):
            nxt = min(len(payload), off + chunk)
            r = self._api(
                "PUT",
                f"/api/upload/sessions/{sid}/chunk",
                headers={"X-Chunk-Offset": str(off), "Content-Type": "application/octet-stream"},
                content=payload[off:nxt],
            )
            r.raise_for_status()
            off = r.json()["uploaded_bytes"]
        out = self._json("POST", f"/api/upload/sessions/{sid}/complete")
        return out

    def _create_observation(self, image_path: str, idx: int = 1) -> dict[str, Any]:
        return self._json(
            "POST",
            "/api/observations",
            json={
                "project_name": "Chaos Project",
                "tower": f"T{(idx % 3) + 1}",
                "floor": f"F{(idx % 8) + 1}",
                "flat": f"A{idx}",
                "room": "Living Room",
                "observation_type": "Plastering",
                "severity": "Major",
                "inspection_status": "Yet to be Confirmed",
                "third_party_status": "Yet to be Confirmed",
                "image_path": image_path,
                "generate_text": True,
            },
        )

    def _generate_report(self, obs_ids: list[int], include_pdf: bool = False) -> dict[str, Any]:
        return self._json(
            "POST",
            "/api/reports/generate",
            json={"observation_ids": obs_ids, "title": "Chaos Deck", "include_pdf": include_pdf},
        )

    def _wait_report_done(self, report_id: int, timeout_s: float = 120.0) -> dict[str, Any]:
        end = time.time() + timeout_s
        while time.time() < end:
            report = self._json("GET", f"/api/reports/{report_id}")
            if report["status"] in {"ready", "failed"}:
                return report
            time.sleep(1.0)
        raise TimeoutError("Report did not finish within timeout")

    def _download_artifact(self, report_id: int, fmt: str) -> Path:
        r = self._api("GET", f"/api/reports/{report_id}/download?format={fmt}")
        r.raise_for_status()
        p = self.out_dir / f"report_{report_id}.{fmt}"
        p.write_bytes(r.content)
        return p

    def _validate_artifacts(self, report_id: int, report: dict[str, Any]) -> dict[str, Any]:
        checks: dict[str, Any] = {}
        if report.get("pptx_path"):
            pptx = self._download_artifact(report_id, "pptx")
            Presentation(str(pptx))
            with zipfile.ZipFile(pptx, "r") as zf:
                checks["pptx_entries"] = len(zf.namelist())
        if report.get("xlsx_path"):
            xlsx = self._download_artifact(report_id, "xlsx")
            wb = load_workbook(filename=str(xlsx), read_only=True)
            checks["xlsx_sheets"] = list(wb.sheetnames)
            wb.close()
        if report.get("pdf_path"):
            pdf = self._download_artifact(report_id, "pdf")
            sig = pdf.read_bytes()[:4]
            checks["pdf_signature_ok"] = sig == b"%PDF"
        return checks

    def run(self) -> dict[str, Any]:
        started = time.time()
        self._start_backend()
        baseline = self._json("GET", "/api/ops/health")
        results: list[ScenarioResult] = []

        def exec_scenario(name: str, fn) -> None:
            t0 = time.time()
            try:
                details = fn()
                results.append(
                    ScenarioResult(
                        name=name,
                        passed=True,
                        duration_ms=int((time.time() - t0) * 1000),
                        details=details or {},
                    )
                )
            except Exception as exc:  # noqa: BLE001
                results.append(
                    ScenarioResult(
                        name=name,
                        passed=False,
                        duration_ms=int((time.time() - t0) * 1000),
                        details={},
                        error=str(exc),
                    )
                )

        def s_resumable_upload() -> dict[str, Any]:
            self._fault_reset()
            payload = self._make_png_bytes()
            out = self._resumable_upload(payload, "resume.png")
            if not out.get("path"):
                raise RuntimeError("Upload finalize returned empty path")
            return {"path": out["path"], "public_id": out.get("public_id", "")}

        def s_dependency_cloudinary_outage() -> dict[str, Any]:
            self._fault_reset()
            self._set_fault("cloudinary", True, mode="outage")
            payload = self._make_png_bytes(color=(12, 140, 80))
            out = self._resumable_upload(payload, "cloud-out.png")
            self._fault_reset()
            # fallback path should still exist
            return {"path": out.get("path", ""), "fallback_local": not str(out.get("path", "")).startswith("http")}

        def s_dependency_llm_outage() -> dict[str, Any]:
            self._fault_reset()
            self._set_fault("llm", True, mode="outage")
            up = self._resumable_upload(self._make_png_bytes(color=(100, 30, 220)), "llm-out.png")
            obs = self._create_observation(up["path"], idx=11)
            self._fault_reset()
            if not obs.get("id"):
                raise RuntimeError("Observation was not created during LLM outage")
            return {"id": obs["id"], "ai_status": obs.get("ai_status")}

        def s_export_restart_recovery() -> dict[str, Any]:
            self._fault_reset()
            up = self._resumable_upload(self._make_png_bytes(color=(120, 120, 40)), "exp-restart.png")
            obs = self._create_observation(up["path"], idx=21)
            accepted = self._generate_report([obs["id"]], include_pdf=False)
            rid = int(accepted["report_id"])
            if self.backend_cmd:
                self._restart_backend()
            report = self._wait_report_done(rid, timeout_s=150.0)
            if report["status"] != "ready":
                raise RuntimeError(f"Recovered report did not become ready: {report['status']}")
            artifacts = self._validate_artifacts(rid, report)
            return {"report_id": rid, "status": report["status"], "artifacts": artifacts}

        def s_export_fault_validation() -> dict[str, Any]:
            self._fault_reset()
            self._set_fault("export", True, mode="fail_after_ppt")
            up = self._resumable_upload(self._make_png_bytes(color=(30, 140, 180)), "exp-fail.png")
            obs = self._create_observation(up["path"], idx=31)
            accepted = self._generate_report([obs["id"]], include_pdf=False)
            rid = int(accepted["report_id"])
            report = self._wait_report_done(rid, timeout_s=120.0)
            self._fault_reset()
            if report["status"] != "failed":
                raise RuntimeError("Expected failed status under injected export fault")
            return {"report_id": rid, "status": report["status"], "error": report.get("error_message", "")}

        def s_scale_and_memory() -> dict[str, Any]:
            self._fault_reset()
            start_health = self._json("GET", "/api/ops/health")
            ids: list[int] = []
            for i in range(1, 21):
                up = self._resumable_upload(self._make_png_bytes(color=(i * 5 % 255, 40, 70)), f"scale-{i}.png")
                obs = self._create_observation(up["path"], idx=100 + i)
                ids.append(obs["id"])
            accepted = self._generate_report(ids[:8], include_pdf=False)
            report = self._wait_report_done(int(accepted["report_id"]), timeout_s=180.0)
            end_health = self._json("GET", "/api/ops/health")
            if report["status"] != "ready":
                raise RuntimeError("Scale report did not complete successfully")
            return {
                "created_observations": len(ids),
                "reports_delta": end_health["counts"]["reports_total"] - start_health["counts"]["reports_total"],
                "observations_delta": end_health["counts"]["observations_total"] - start_health["counts"]["observations_total"],
            }

        exec_scenario("resumable_upload_resume", s_resumable_upload)
        exec_scenario("dependency_cloudinary_outage_fallback", s_dependency_cloudinary_outage)
        exec_scenario("dependency_llm_outage_fallback", s_dependency_llm_outage)
        exec_scenario("backend_restart_export_recovery", s_export_restart_recovery)
        exec_scenario("export_fault_state_validation", s_export_fault_validation)
        exec_scenario("scale_and_memory_smoke", s_scale_and_memory)

        final_health = self._json("GET", "/api/ops/health")
        passed = sum(1 for r in results if r.passed)
        total = len(results)
        resilience_score = round((passed / total) * 100, 2) if total else 0.0
        summary = {
            "started_at": utc_now_iso(),
            "duration_ms": int((time.time() - started) * 1000),
            "base_url": self.base_url,
            "baseline": baseline,
            "final_health": final_health,
            "scenario_results": [r.as_dict() for r in results],
            "scores": {
                "resilience_score": resilience_score,
                "stability_score": resilience_score,
                "pass_count": passed,
                "fail_count": total - passed,
            },
        }
        return summary


def write_reports(summary: dict[str, Any], out_dir: Path) -> None:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    json_path = out_dir / f"chaos-report-{stamp}.json"
    md_path = out_dir / f"chaos-report-{stamp}.md"
    json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    lines = [
        "# SiteLens Chaos Report",
        "",
        f"- Generated: `{summary['started_at']}`",
        f"- Duration: `{summary['duration_ms']} ms`",
        f"- Resilience score: `{summary['scores']['resilience_score']}`",
        f"- Stability score: `{summary['scores']['stability_score']}`",
        f"- Pass: `{summary['scores']['pass_count']}`",
        f"- Fail: `{summary['scores']['fail_count']}`",
        "",
        "## Scenario Matrix",
        "",
    ]
    for row in summary["scenario_results"]:
        status = "PASS" if row["passed"] else "FAIL"
        lines.append(f"- `{status}` `{row['name']}` in `{row['duration_ms']} ms`")
        if row.get("error"):
            lines.append(f"  - error: `{row['error']}`")
    lines.extend(
        [
            "",
            "## Baseline Counts",
            "",
            f"- `{summary['baseline']['counts']}`",
            "",
            "## Final Counts",
            "",
            f"- `{summary['final_health']['counts']}`",
            "",
            "## Breakers",
            "",
            f"- `{summary['final_health'].get('breakers', [])}`",
            "",
            f"JSON report: `{json_path.name}`",
        ]
    )
    md_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {json_path}")
    print(f"Wrote {md_path}")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="SiteLens automated chaos runner")
    p.add_argument("--base-url", default="http://127.0.0.1:8080", help="Backend base URL")
    p.add_argument("--out-dir", default="chaos/results", help="Result output directory")
    p.add_argument(
        "--backend-cmd",
        default="",
        help="Optional command to launch backend; enables restart scenario (example: uvicorn app.main:app --host 127.0.0.1 --port 8080)",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()
    out_dir = Path(args.out_dir)
    runner = ChaosRunner(
        base_url=args.base_url,
        out_dir=out_dir,
        backend_cmd=args.backend_cmd or None,
    )
    try:
        summary = runner.run()
        write_reports(summary, out_dir)
        return 0 if summary["scores"]["fail_count"] == 0 else 2
    finally:
        runner.close()


if __name__ == "__main__":
    raise SystemExit(main())
